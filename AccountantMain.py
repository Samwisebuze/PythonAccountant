################
# Python Accountant
# By: Samuel Buzas & Elizabeth Sheetz
# Version 0.1
#################

from PyAcc_Transactions import *
from PyAcc_Account import *
from openpyxl import Workbook
import time

import os
def cls():
    os.system('cls' if os.name=='nt' else 'clear')


def main():
	wb = Workbook()
	account = Account()
	cls()
	print("\t\t Welcome to Python Accountant!\n")

	while True:
		cls()
		print("What Would you like to do today? \n")
		#Menu
		print("You are Curently Editing the " + "/ " + account.account_Name +" /" +" Account in the " + account.workbook_name +" workbook.\n\n")
		print("\t\t Main Menu \n")
		print("\t 1 : \t Create a New Budget \n")
		print("\t 2 : \t Create a New Account (Must have an existing Workbook) \n")
		print("\t 3 : \t Select a Budget and an account to work with.\n")
		print("\t\t 3a : \t Enter a New Transaction \n")
		print("\t\t 3b : \t Edit an Existing Transaction (Must know Transaction number (Row Number)) \n")
		print("\t 8 : \t Exit Python Accountant  \n")

		option = input("Please Input an option: ")
		option = str(option)
		
		# Cases
		# Create a New Budget
		if option == '1':
			print("Let's Create a file.\n")
			filename = input("Enter a filename: ")
			worksheet = input("Enter a worksheet name: ")
			create_workbook(filename,worksheet)


		# Create a New Account
		elif option == '2':
			print("Let's Create a new Account.\n")
			filename = input("Enter an existing filename  (No need to include the file extension): ")
			account_name =input("Enter a Account Name name: ")
			init_balance = input("Enter the initial balance: ")
			account = Account(account_name,filename,init_balance)
			transaction = Transaction(filename,account_name)
			account.create_new_Account()
			

		# Select a Budget and an account to work with
		elif option == '3':
			print("What workbook would you like to work with?\n")
			filename = input("Enter A Filename: ")
			wb = load_workbook(str(filename) +'.xlsx')

			print("Which Account would you like to work on?\n")
			print(wb.sheetnames)
			account_name = input("Please Enter an Account: ")
			transaction = Transaction(filename,account_name)
			account = Account(account_name,filename)

		# Enter a New Transaction
		elif option == '3a':
			print("Lets add an new Transaction.\n")
			# year,month,day,category,amount,checknum,desc
			# Date
			print("What is the Date of the Transaction?\n")
			year = input("Enter the Year: ")
			month = input("Enter the Month: ")
			day = input("Enter the Day: ")

			print("How much was the transaction?\n Does it have a category?\n")
			amount = input("Enter the amount (Use negative for withdrawals and purchases): ")
			category = input("Enter the Category: ")

			checknum = input("If you used a check, please enter the check number here (Otherwise leave blank): ")
			print("Leave a small description of the transaction here:\n")
			description = input()

			transaction.createTransaction(int(year),int(month),int(day),category,float(amount),checknum,description)
			account.balance_update()


		#Edit an Existing Transaction
		elif option == '3b':
			exit = False
			#Sub-Menu
			while exit == False:
				cls()
				print("\t\t Welcome to Python Accountant!\n")
				print("What Would you like to do today? \n")
				#Menu
				print("You are Curently Editing the " + account.account_Name + " Account in the " + account.workbook_name +" workbook.\n\n")
				print("\t\t Transaction Edit Menu \n")
				print("\t 1: \t Update the Date\n")
				print("\t 2: \t Update the Category\n")
				print("\t 3: \t Update the Amount\n")
				print("\t 4: \t Update the Check Number\n")
				print("\t 5: \t Update the Description\n")
				print("\t 6: \t Back to Main Menu\n")

				suboption = input("Please Input an option: ")
				suboption = str(suboption)

				#Date
				if suboption == '1':
					entry = input("Enter the transaction number: ")
					print("What is the new date of the transaction?\n")
					year = input("Enter the Year: ")
					month = input("Enter the Month: ")
					day = input("Enter the Day: ")
					transaction.updateTransaction_Date(entry,int(year),int(month),int(day))

				#Category
				if suboption == '2':
					entry = input("Enter the transaction number: ")
					category = input("Enter the new Category: ")
					transaction.updateTransaction_Category(entry,category)

				#Amount
				if suboption == '3':
					entry = input("Enter the transaction number: ")
					amount = input("Enter the new amount (Use negative for withdrawls and purchases): ")
					transaction.updateTransaction_amount(entry,amount)
					account.balance_update()

				#Check Number
				if suboption == '4':
					entry = input("Enter the transaction number: ")
					checknum = input("Enter the new Check Number: ")
					transaction.updateTransaction_checknum(entry,checknum)
		

				#Description
				if suboption == '5':
					entry = input("Enter the transaction number: ")
					print("Enter the new description of the transaction here:\n")
					description = input()
					transaction.updateTransaction_desc(entry, description)
		

				if suboption == '6':
					exit = True




		elif option == '8':
			print("Thank you for chosing Python Accountant for you budgeting needs!\n\n\n")
			print("( *A*,)/\n\n\n")
			time.sleep(2)
			import sys
			sys.exit(1)

		else:
			print("Please make a vaild selection from the Menu \n\n" )
			time.sleep(3)


if __name__ == "__main__":
	main()


