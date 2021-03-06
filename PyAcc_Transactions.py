#########
<<<<<<< HEAD
# Python Accounatn Header
# Desc: Contains Any functions or shortcut snippets used to sumplify main
=======
# Python Accountant Header 
# Desc: Contains Any functions or shortcut snippets used to sumplify main, including Transaction Class Definition
>>>>>>> origin/master
# Created By: Samuel Buzas & Elizabeth Sheetz
#########
import datetime

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

#Constant
wb = Workbook()

# Create Workbook
<<<<<<< HEAD
def create_workbook(filename = "MyFirstBudget" + '.xlsx', sheet_name = "SimpleBudget"):
=======
def create_workbook(filename = "MyFirstBudget", sheet_name = "SimpleBudget"):
	#Creates a new workbooks (i.e. a new excel file)
	#Also creates a new workseet
	dest_filename = filename + '.xlsx'
>>>>>>> origin/master
	wb.active.title = sheet_name
	wb.active['A1'] = 'Date'
	wb.active.column_dimensions['A'].width = 17
	wb.active['B1'] = 'Category'
	wb.active.column_dimensions['B'].width = 15
	wb.active['C1'] = 'Amount'
	wb.active.column_dimensions['C'].width = 10
	wb.active['D1'] = 'Check Number'
	wb.active.column_dimensions['D'].width = 17
	wb.active['E1'] = 'Description'
	wb.active.column_dimensions['E'].width = 50
	wb.geuss_types = True
	wb.save(filename)

<<<<<<< HEAD
# Create WorkSheet
def create_worksheet(sheet_name = "Trackin Sheet", filename = "MyFirstBudget"+ '.xlsx'):

	wb = load_workbook(filename)
	wb.create_sheet(sheet_name)
	wb.save(filename)

# Create Blank Graph
def create_table(filename = "MyFirstBudget"+'.xlsx', sheet_name = "SimpleBudget"):
	wb.active.title = sheet_name
	wb.save(filename)
=======
# Create WorkSheet (did not use) (replaced with create_new_Account in PyAcc_Accountpy)
# def create_worksheet(sheet_name = "Tracking Sheet", filename = "MyFirstBudget"):
# 	dest_filename = filename + '.xlsx'
# 	wb = load_workbook(dest_filename)
# 	wb.create_sheet(sheet_name)
# 	wb.save(dest_filename)

# Create Blank Graph (did not use)
# def create_table(filename = "MyFirstBudget", sheet_name = "SimpleBudget"):
# 	dest_filename = filename + '.xlsx'
# 	wb.active.title = sheet_name
# 	wb.save(dest_filename)
>>>>>>> origin/master



class Transaction:

<<<<<<< HEAD

	def __init__(self, workbook = 'Default'+'.xlsx',account_name = 'Default'):
		self.workbook = workbook
=======
	def __init__(self, workbook = 'Default' ,account_name = 'Default'):
		self.workbook = workbook + '.xlsx'
>>>>>>> origin/master
		self.account_name = str(account_name)
		# excel isnt 0 based and row 1 is for titles so first entry = 2
		self.entryNumber = 2

	# Create Entry
	def createTransaction(self, year = 1996 , month = 12,  day = 12, category = 'None' , amount = 0.0, checknum = 'None' , desc ='Ex.This was a purchase'):
		# Load budget workbook to edit
		wb = load_workbook(self.workbook)
		sheet = wb[self.account_name]
		# check that you're not overwriting previous entries 
		# i.e. find appropriate entrynumber so you add entry at end of list of previously entered transactions
		# Always start at entry 2, to prevent massive gaps in the sheet
		self.entryNumber = 2
		while sheet['A'+ str(self.entryNumber)].value != None:
			self.entryNumber += 1

		sheet['A' + str(self.entryNumber)] = datetime.datetime(year, month, day)
		sheet['A' + str(self.entryNumber)].number_format
		wb.geuss_types= True
		sheet['B' + str(self.entryNumber)] = category
		sheet['C' + str(self.entryNumber)] = amount
		sheet['D' + str(self.entryNumber)] = checknum
		sheet['E' + str(self.entryNumber)] = desc

		wb.save(self.workbook)
		self.entryNumber += 1

	def checkEntries(self,filename,account_name):
		#returns first line in filename.account that has no transaction entered
		wb = load_workbook(filename)
		sheet = wb[account_name]
		self.entryNumber = 2
		while sheet['A'+ str(self.entryNumber)].value != None:
			self.entryNumber += 1
		return self.entryNumber



	# Edit Entries
	# In future would like to consolidate these
		# eg. updateTransaction_bytype(self, entryNumber, datatype, newvalue) would do all the stuff they all do, and then some sort of way to customize for specific datatypes
	def updateTransaction_Date(self, entryNumber, year, month, day):
		wb = load_workbook(self.workbook)
		sheet = wb[self.account_name]
		self.entryNumber = entryNumber
		# Update the description
		sheet['A' + str(self.entryNumber)] = datetime.datetime(year, month, day)
		sheet['A' + str(self.entryNumber)].number_format
		wb.geuss_types = True

		wb.save(self.workbook)

	def updateTransaction_Category(self, entryNumber, category):
		wb = load_workbook(self.workbook)
		sheet = wb[self.account_name]
		self.entryNumber = entryNumber
		# Update the description
		sheet['B' + str(self.entryNumber)] = category

		wb.save(self.workbook)

	def updateTransaction_amount(self, entryNumber, amount):
		wb = load_workbook(self.workbook)
		sheet = wb[self.account_name]
		self.entryNumber = entryNumber
		# Update the description
		sheet['C' + str(self.entryNumber)] = amount

		wb.save(self.workbook)

	def updateTransaction_checknum(self,entryNumber, checknum):
		wb = load_workbook(self.workbook)
		sheet = wb[self.account_name]
		self.entryNumber = entryNumber
		# Update the description
		sheet['D' + str(self.entryNumber)] = checknum

		wb.save(self.workbook)

	def updateTransaction_desc(self, entryNumber, desc):
		wb = load_workbook(self.workbook)
		sheet = wb[self.account_name]
		self.entryNumber = entryNumber
		# Update the description
		sheet['E' + str(self.entryNumber)] = desc

<<<<<<< HEAD
		wb.save(self.workbook)
=======
		wb.save(self.workbook)
>>>>>>> origin/master
