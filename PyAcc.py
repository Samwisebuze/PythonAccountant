#########
# Python Accounatn Header 
# Desc: Contains Any functions or shortcut snippets used to sumplify main
# Created By: Samuel Buzas & Elizabeth Sheetz
#########
import datetime
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

#Constant
wb = Workbook()

# Create Workbook
def create_workbook(filename = "MyFirstBudget", sheet_name = "SimpleBudget"):
	dest_filename = filename + '.xlsx'
	wb.active.title = sheet_name
	wb.active['A1'] = 'Date'
	wb.active.column_dimensions['A'].width = 10
	wb.active['B1'] = 'Category'
	wb.active.column_dimensions['B'].width = 10
	wb.active['C1'] = 'Amount'
	wb.active['D1'] = 'Check Number'
	wb.active['E1'] = 'Description'
	wb.active.column_dimensions['E'].width = 30
	wb.geuss_types = True
	wb.save(filename = dest_filename)

# Create WorkSheet
def create_worksheet(sheet_name = "Trackin Sheet", filename = "MyFirstBudget"):
	dest_filename = filename + '.xlsx'
	wb.create_sheet(sheet_name)
	wb.save(filename = dest_filename)

# Create Blank Graph
def create_table(filename = "MyFirstBudget", sheet_name = "SimpleBudget"):
	dest_filename = filename + '.xlsx'
	wb.active.title = sheet_name
	wb.save(filename = dest_filename)



class Transaction:
	
	

	def __init__(self, Day = 12, Month =12, Year=12 , category = 'None' , amount = 0.0, checknum = 'None' , desc ='Ex. This was a purcahse'):
		# Date format is Month/ Day / Year
		self.Day = Day
		self.Month = Month
		self.Year = Year
		#Category is used to lable puchases (i.e: Food, Entertainment, Gas, Travel,...etc)
		self.category = category
		# The Amount of the transaction
		self.amount = amount
		# The Check Number of the transaction, if needed
		self.checknum = checknum
		# Short Description of the purpose of the transaction, further defines the transaction with help of category
		self.desc = desc
		# excel isnt 0 based and row 1 is for titles
		self.entryNumber = 2
	
	# Create Entry
	def createTransaction(self, day= 12, month= 12, year= 12 , category = 'None' , amount = 0.0, checknum = 'None' , desc ='Ex.This was a purcahse'):
		sheet = wb.active
		sheet['A' + str(self.entryNumber)] = datetime.datetime(year, month, day)
		sheet['A' + str(self.entryNumber)].number_format
		wb.geuss_types= True
		sheet['B' + str(self.entryNumber)] = category
		sheet['C' + str(self.entryNumber)] = amount
		sheet['D' + str(self.entryNumber)] = checknum
		sheet['E' + str(self.entryNumber)] = desc

		wb.save('MyFirstBudget.xlsx')
		self.entryNumber += 1

# Edit Entry


